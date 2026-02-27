#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
PDF-to-PDF comparison with fuzzy block alignment and Excel reporting.

Usage:
  python pdf_comparison_tool.py old.pdf new.pdf -o diff.xlsx --min-sim 0.70

Notes:
- Designed for text-based PDFs (not scanned images).
- Compares paragraph-like blocks, aligns them with fuzzy matching, reports differences.
- Includes GUI for easy file selection
"""

import re
import sys
import argparse
from dataclasses import dataclass
from typing import List, Dict, Set, Tuple, Optional
from difflib import SequenceMatcher
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import os

try:
    import fitz  # PyMuPDF
except ImportError:
    print("Missing dependency: pymupdf. Install via: pip install pymupdf", file=sys.stderr)
    raise

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
except ImportError:
    print("Missing dependency: openpyxl. Install via: pip install openpyxl", file=sys.stderr)
    raise


@dataclass
class Block:
    block_id: str
    page: int
    text_raw: str
    text_norm: str
    tokens: Set[str]


def normalize_text(s: str) -> str:
    """Normalize text by removing hyphenation and standardizing whitespace"""
    # Remove hyphenation across line breaks: "Auto-\nmatisierung" -> "Automatisierung"
    s = re.sub(r"(\w)-\n(\w)", r"\1\2", s)
    # Normalize line breaks
    s = s.replace("\r\n", "\n").replace("\r", "\n")
    # Collapse whitespace
    s = re.sub(r"[ \t]+", " ", s)
    # Normalize newlines: keep paragraph breaks, but avoid multiple blank lines
    s = re.sub(r"\n[ \t]*\n+", "\n\n", s)
    # Trim
    s = s.strip()
    return s


def tokenize(s: str) -> Set[str]:
    """Extract tokens for matching (words with 4+ characters)"""
    words = re.findall(r"[A-Za-zÄÖÜäöüß0-9]+", s.lower())
    # drop very short tokens (noise)
    return {w for w in words if len(w) >= 4}


def extract_pages_text(pdf_path: str) -> List[str]:
    """Extract text from all pages of a PDF"""
    doc = fitz.open(pdf_path)
    pages = []
    for i, page in enumerate(doc, start=1):
        txt = page.get_text("text") or ""
        pages.append(txt)
    doc.close()
    return pages


def detect_repeating_lines(pages_text: List[str], threshold_ratio: float = 0.6) -> Set[str]:
    """
    Heuristic for header/footer removal:
    Count normalized lines that repeat on many pages.
    """
    line_counts: Dict[str, int] = {}
    total_pages = len(pages_text)
    for t in pages_text:
        lines = [re.sub(r"\s+", " ", ln.strip()) for ln in t.splitlines()]
        # ignore empty lines
        lines = [ln for ln in lines if ln]
        # count unique per page to avoid overcounting
        for ln in set(lines):
            if 3 <= len(ln) <= 120:
                line_counts[ln] = line_counts.get(ln, 0) + 1

    cutoff = max(2, int(total_pages * threshold_ratio))
    repeating = {ln for ln, c in line_counts.items() if c >= cutoff}
    return repeating


def strip_headers_footers(page_text: str, repeating_lines: Set[str]) -> str:
    """Remove repeating headers and footers from page text"""
    lines = page_text.splitlines()
    norm_lines = [re.sub(r"\s+", " ", ln.strip()) for ln in lines]

    # Remove repeating lines only if they appear near top/bottom (reduce false positives)
    top_k = 4
    bottom_k = 4
    to_remove_idx = set()

    for idx in range(min(top_k, len(norm_lines))):
        if norm_lines[idx] in repeating_lines:
            to_remove_idx.add(idx)

    for idx in range(max(0, len(norm_lines) - bottom_k), len(norm_lines)):
        if norm_lines[idx] in repeating_lines:
            to_remove_idx.add(idx)

    kept = [ln for i, ln in enumerate(lines) if i not in to_remove_idx]
    return "\n".join(kept)


def segment_into_blocks(pages_text: List[str], label_prefix: str) -> List[Block]:
    """
    Build paragraph-like blocks per page using blank lines as separators,
    plus a safety split for very long blocks.
    """
    repeating = detect_repeating_lines(pages_text)
    blocks: List[Block] = []

    for pageno, raw in enumerate(pages_text, start=1):
        raw2 = strip_headers_footers(raw, repeating)
        raw2 = normalize_text(raw2)

        # Split by blank lines into paragraphs
        parts = [p.strip() for p in raw2.split("\n\n") if p.strip()]

        # Safety: split very long paragraphs into chunks
        chunked: List[str] = []
        for part in parts:
            if len(part) <= 1400:
                chunked.append(part)
            else:
                # split on sentence-ish boundaries
                sentences = re.split(r"(?<=[\.\!\?])\s+", part)
                buf = []
                buf_len = 0
                for s in sentences:
                    if not s:
                        continue
                    if buf_len + len(s) + 1 > 1200 and buf:
                        chunked.append(" ".join(buf).strip())
                        buf = [s]
                        buf_len = len(s)
                    else:
                        buf.append(s)
                        buf_len += len(s) + 1
                if buf:
                    chunked.append(" ".join(buf).strip())

        for idx, txt in enumerate(chunked, start=1):
            norm = normalize_text(txt)
            tok = tokenize(norm)
            bid = f"{label_prefix}-p{pageno:03d}-b{idx:03d}"
            blocks.append(Block(block_id=bid, page=pageno, text_raw=txt, text_norm=norm, tokens=tok))

    return blocks


def jaccard(a: Set[str], b: Set[str]) -> float:
    """Calculate Jaccard similarity between two sets"""
    if not a and not b:
        return 1.0
    if not a or not b:
        return 0.0
    inter = len(a & b)
    union = len(a | b)
    return inter / union if union else 0.0


def similarity(a: Block, b: Block) -> float:
    """Calculate similarity score between two blocks"""
    # Combine token Jaccard with string similarity
    seq = SequenceMatcher(None, a.text_norm, b.text_norm).ratio()
    jac = jaccard(a.tokens, b.tokens)
    return 0.6 * seq + 0.4 * jac


def build_token_index(blocks: List[Block]) -> Tuple[Dict[str, Set[int]], Dict[str, int]]:
    """Build inverted index and document frequency for tokens"""
    idx: Dict[str, Set[int]] = {}
    df: Dict[str, int] = {}
    for i, bl in enumerate(blocks):
        for t in bl.tokens:
            idx.setdefault(t, set()).add(i)
    for t, s in idx.items():
        df[t] = len(s)
    return idx, df


def pick_candidates(new_block: Block, old_index: Dict[str, Set[int]], df: Dict[str, int], total_old: int) -> Set[int]:
    """
    Candidate retrieval: pick rare tokens in new_block and union the blocks that contain them.
    Avoid very common tokens.
    """
    if not new_block.tokens:
        return set(range(total_old))

    # sort tokens by document frequency (rare first)
    toks = sorted(new_block.tokens, key=lambda t: df.get(t, total_old + 1))

    # take a handful of rare tokens
    chosen = []
    for t in toks:
        # skip very common tokens
        if df.get(t, 0) > max(10, int(0.25 * total_old)):
            continue
        chosen.append(t)
        if len(chosen) >= 8:
            break

    if not chosen:
        return set(range(total_old))

    cand: Set[int] = set()
    for t in chosen:
        cand |= old_index.get(t, set())

    return cand if cand else set(range(total_old))


def word_level_diff(old_text: str, new_text: str, max_chars: int = 5000) -> str:
    """Generate word-level diff with markup"""
    old_words = old_text.split()
    new_words = new_text.split()
    sm = SequenceMatcher(None, old_words, new_words)
    out = []
    for tag, i1, i2, j1, j2 in sm.get_opcodes():
        if tag == "equal":
            out.extend(old_words[i1:i2])
        elif tag == "delete":
            out.append("[-" + " ".join(old_words[i1:i2]) + "-]")
        elif tag == "insert":
            out.append("{+" + " ".join(new_words[j1:j2]) + "+}")
        elif tag == "replace":
            out.append("[-" + " ".join(old_words[i1:i2]) + "-]")
            out.append("{+" + " ".join(new_words[j1:j2]) + "+}")

        if sum(len(x) for x in out) > max_chars:
            out.append("…")
            break
    return " ".join(out)


def align_blocks(old_blocks: List[Block], new_blocks: List[Block], min_sim: float) -> Tuple[List[Tuple[Optional[int], Optional[int], float]], Set[int], Set[int]]:
    """
    Align blocks between old and new PDFs
    Returns:
      pairs: list of (old_idx or None, new_idx or None, sim)
      unmatched_old, unmatched_new
    """
    old_index, df = build_token_index(old_blocks)
    unmatched_old = set(range(len(old_blocks)))
    pairs: List[Tuple[Optional[int], Optional[int], float]] = []

    for j, nb in enumerate(new_blocks):
        candidates = pick_candidates(nb, old_index, df, len(old_blocks))
        candidates &= unmatched_old

        best_i = None
        best_s = -1.0

        for i in candidates:
            s = similarity(old_blocks[i], nb)
            if s > best_s:
                best_s = s
                best_i = i

        if best_i is not None and best_s >= min_sim:
            pairs.append((best_i, j, best_s))
            unmatched_old.remove(best_i)
        else:
            pairs.append((None, j, 0.0))

    unmatched_new = set(range(len(new_blocks)))
    for oi, nj, _ in pairs:
        if nj is not None:
            unmatched_new.discard(nj)

    return pairs, unmatched_old, unmatched_new


def write_excel(
    out_path: str,
    old_pdf: str,
    new_pdf: str,
    old_blocks: List[Block],
    new_blocks: List[Block],
    pairs: List[Tuple[Optional[int], Optional[int], float]],
    unmatched_old: Set[int],
    min_sim: float,
) -> None:
    """Write comparison results to Excel file"""
    wb = Workbook()

    # Summary sheet
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum["A1"] = "PDF Diff Summary"
    ws_sum["A1"].font = Font(bold=True, size=14)
    ws_sum["A3"] = "Old PDF"
    ws_sum["B3"] = old_pdf
    ws_sum["A4"] = "New PDF"
    ws_sum["B4"] = new_pdf
    ws_sum["A6"] = "Old blocks"
    ws_sum["B6"] = len(old_blocks)
    ws_sum["A7"] = "New blocks"
    ws_sum["B7"] = len(new_blocks)
    ws_sum["A8"] = "Min similarity"
    ws_sum["B8"] = min_sim

    # Changes sheet
    ws = wb.create_sheet("Changes")
    headers = [
        "change_type", "similarity",
        "old_page", "new_page",
        "old_block_id", "new_block_id",
        "old_snippet", "new_snippet",
        "diff_markup"
    ]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        ws.cell(row=1, column=c).font = Font(bold=True)

    def snippet(s: str, n: int = 800) -> str:
        s = s.strip()
        return s if len(s) <= n else s[:n] + "…"

    changed_count = 0
    added_count = 0
    removed_count = 0
    unchanged_count = 0
    modified_count = 0

    # matched or added
    for oi, nj, sim in pairs:
        nb = new_blocks[nj] if nj is not None else None

        if oi is None and nb is not None:
            added_count += 1
            ws.append([
                "added", 0.0,
                None, nb.page,
                None, nb.block_id,
                "", snippet(nb.text_raw),
                "{+NEW BLOCK+}"
            ])
            changed_count += 1
            continue

        if oi is not None and nb is not None:
            ob = old_blocks[oi]
            # Determine unchanged vs modified
            if ob.text_norm == nb.text_norm:
                unchanged_count += 1
                continue

            modified_count += 1
            diff = word_level_diff(ob.text_norm, nb.text_norm)
            ws.append([
                "modified", round(sim, 4),
                ob.page, nb.page,
                ob.block_id, nb.block_id,
                snippet(ob.text_raw), snippet(nb.text_raw),
                diff
            ])
            changed_count += 1

    # removed
    for oi in sorted(unmatched_old):
        ob = old_blocks[oi]
        removed_count += 1
        ws.append([
            "removed", 0.0,
            ob.page, None,
            ob.block_id, None,
            snippet(ob.text_raw), "",
            "[-REMOVED BLOCK-]"
        ])
        changed_count += 1

    # Put counts into summary
    ws_sum["A10"] = "Modified"
    ws_sum["B10"] = modified_count
    ws_sum["A11"] = "Added"
    ws_sum["B11"] = added_count
    ws_sum["A12"] = "Removed"
    ws_sum["B12"] = removed_count
    ws_sum["A13"] = "Unchanged (skipped rows)"
    ws_sum["B13"] = unchanged_count
    ws_sum["A14"] = "Reported rows"
    ws_sum["B14"] = changed_count

    # Formatting
    ws.freeze_panes = "A2"
    wrap = Alignment(wrap_text=True, vertical="top")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.alignment = wrap

    # Column widths
    col_widths = {
        "A": 12, "B": 10, "C": 9, "D": 9, "E": 18, "F": 18, "G": 45, "H": 45, "I": 60
    }
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    wb.save(out_path)


def run_pdf_comparison_gui():
    """Start GUI for PDF comparison"""
    root = tk.Tk()
    root.title("PDF Vergleichstool")
    root.geometry("650x450")
    
    # Variables
    old_pdf_path = tk.StringVar()
    new_pdf_path = tk.StringVar()
    output_path = tk.StringVar()
    min_similarity = tk.DoubleVar(value=0.70)
    
    # Title
    title_label = tk.Label(root, text="PDF Vergleichstool", font=("Arial", 16, "bold"))
    title_label.pack(pady=15)
    
    # File selection frame
    file_frame = ttk.Frame(root, padding=10)
    file_frame.pack(fill=tk.BOTH, expand=True, padx=20)
    
    # Old PDF
    ttk.Label(file_frame, text="Alte PDF-Datei:", font=("Arial", 10, "bold")).grid(row=0, column=0, sticky=tk.W, pady=5)
    old_pdf_entry = ttk.Entry(file_frame, textvariable=old_pdf_path, width=50)
    old_pdf_entry.grid(row=1, column=0, pady=5, padx=(0, 5))
    
    def browse_old_pdf():
        filename = filedialog.askopenfilename(
            title="Alte PDF-Datei auswählen",
            filetypes=[("PDF Dateien", "*.pdf"), ("Alle Dateien", "*.*")]
        )
        if filename:
            old_pdf_path.set(filename)
    
    ttk.Button(file_frame, text="Durchsuchen...", command=browse_old_pdf).grid(row=1, column=1, pady=5)
    
    # New PDF
    ttk.Label(file_frame, text="Neue PDF-Datei:", font=("Arial", 10, "bold")).grid(row=2, column=0, sticky=tk.W, pady=5)
    new_pdf_entry = ttk.Entry(file_frame, textvariable=new_pdf_path, width=50)
    new_pdf_entry.grid(row=3, column=0, pady=5, padx=(0, 5))
    
    def browse_new_pdf():
        filename = filedialog.askopenfilename(
            title="Neue PDF-Datei auswählen",
            filetypes=[("PDF Dateien", "*.pdf"), ("Alle Dateien", "*.*")]
        )
        if filename:
            new_pdf_path.set(filename)
    
    ttk.Button(file_frame, text="Durchsuchen...", command=browse_new_pdf).grid(row=3, column=1, pady=5)
    
    # Output Excel
    ttk.Label(file_frame, text="Ausgabe Excel-Datei:", font=("Arial", 10, "bold")).grid(row=4, column=0, sticky=tk.W, pady=5)
    output_entry = ttk.Entry(file_frame, textvariable=output_path, width=50)
    output_entry.grid(row=5, column=0, pady=5, padx=(0, 5))
    
    def browse_output():
        filename = filedialog.asksaveasfilename(
            title="Ausgabedatei speichern als",
            defaultextension=".xlsx",
            filetypes=[("Excel Dateien", "*.xlsx"), ("Alle Dateien", "*.*")]
        )
        if filename:
            output_path.set(filename)
    
    ttk.Button(file_frame, text="Durchsuchen...", command=browse_output).grid(row=5, column=1, pady=5)
    
    # Similarity threshold
    ttk.Label(file_frame, text="Min. Ähnlichkeit (0.0 - 1.0):", font=("Arial", 10)).grid(row=6, column=0, sticky=tk.W, pady=10)
    similarity_spinbox = ttk.Spinbox(file_frame, from_=0.0, to=1.0, increment=0.05, textvariable=min_similarity, width=10)
    similarity_spinbox.grid(row=6, column=1, pady=10, sticky=tk.W)
    
    # Progressbar
    progress = ttk.Progressbar(root, mode='indeterminate', length=550)
    progress.pack(pady=10)
    
    # Status Label
    status_label = tk.Label(root, text="Bereit - Wähle PDF-Dateien aus", fg="blue", font=("Arial", 9))
    status_label.pack(pady=5)
    
    def start_comparison():
        old_pdf = old_pdf_path.get()
        new_pdf = new_pdf_path.get()
        output = output_path.get()
        min_sim = min_similarity.get()
        
        # Validation
        if not old_pdf or not os.path.exists(old_pdf):
            messagebox.showerror("Fehler", "Bitte wähle eine gültige alte PDF-Datei aus!")
            return
        
        if not new_pdf or not os.path.exists(new_pdf):
            messagebox.showerror("Fehler", "Bitte wähle eine gültige neue PDF-Datei aus!")
            return
        
        if not output:
            messagebox.showerror("Fehler", "Bitte wähle einen Ausgabepfad für die Excel-Datei!")
            return
        
        try:
            progress.start(10)
            status_label.config(text="[1/4] Extrahiere Text aus alter PDF...", fg="orange")
            root.update()
            old_pages = extract_pages_text(old_pdf)
            
            status_label.config(text="[2/4] Extrahiere Text aus neuer PDF...", fg="orange")
            root.update()
            new_pages = extract_pages_text(new_pdf)
            
            status_label.config(text="[3/4] Segmentiere Textblöcke...", fg="orange")
            root.update()
            old_blocks = segment_into_blocks(old_pages, label_prefix="OLD")
            new_blocks = segment_into_blocks(new_pages, label_prefix="NEW")
            
            status_label.config(text="[4/4] Vergleiche und erstelle Excel-Report...", fg="orange")
            root.update()
            pairs, unmatched_old, unmatched_new = align_blocks(old_blocks, new_blocks, min_sim=min_sim)
            
            write_excel(
                out_path=output,
                old_pdf=old_pdf,
                new_pdf=new_pdf,
                old_blocks=old_blocks,
                new_blocks=new_blocks,
                pairs=pairs,
                unmatched_old=unmatched_old,
                min_sim=min_sim,
            )
            
            progress.stop()
            status_label.config(text="✓ Erfolgreich abgeschlossen!", fg="green")
            
            messagebox.showinfo(
                "Erfolg", 
                f"PDF-Vergleich abgeschlossen!\n\nErgebnis: {output}\n\n"
                f"Alte Blöcke: {len(old_blocks)}\n"
                f"Neue Blöcke: {len(new_blocks)}\n"
                f"Geändert: {len(pairs)}\n"
                f"Entfernt: {len(unmatched_old)}"
            )
            
        except Exception as e:
            progress.stop()
            status_label.config(text="✗ Fehler aufgetreten!", fg="red")
            messagebox.showerror("Fehler", f"Ein Fehler ist aufgetreten:\n\n{str(e)}")
    
    # Buttons
    button_frame = ttk.Frame(root)
    button_frame.pack(pady=15)
    
    start_button = ttk.Button(button_frame, text="▶ Vergleich starten", command=start_comparison)
    start_button.pack(side=tk.LEFT, padx=5)
    
    close_button = ttk.Button(button_frame, text="Schließen", command=root.destroy)
    close_button.pack(side=tk.LEFT, padx=5)
    
    root.mainloop()


def main() -> int:
    """Main function for command-line usage"""
    ap = argparse.ArgumentParser(description="Compare two PDFs and export changes to Excel.")
    ap.add_argument("old_pdf", nargs='?', help="Path to old PDF")
    ap.add_argument("new_pdf", nargs='?', help="Path to new PDF")
    ap.add_argument("-o", "--output", default="pdf_diff.xlsx", help="Output Excel file (default: pdf_diff.xlsx)")
    ap.add_argument("--min-sim", type=float, default=0.70, help="Min similarity for block matching (default: 0.70)")
    ap.add_argument("--gui", action="store_true", help="Launch GUI instead of CLI")
    args = ap.parse_args()

    # If no arguments or --gui flag, launch GUI
    if args.gui or (not args.old_pdf and not args.new_pdf):
        run_pdf_comparison_gui()
        return 0

    # CLI mode
    if not args.old_pdf or not args.new_pdf:
        ap.print_help()
        return 1

    print(f"[1/4] Extracting text: {args.old_pdf}")
    old_pages = extract_pages_text(args.old_pdf)
    print(f"[2/4] Extracting text: {args.new_pdf}")
    new_pages = extract_pages_text(args.new_pdf)

    print("[3/4] Segmenting into blocks …")
    old_blocks = segment_into_blocks(old_pages, label_prefix="OLD")
    new_blocks = segment_into_blocks(new_pages, label_prefix="NEW")

    print("[4/4] Aligning + exporting …")
    pairs, unmatched_old, unmatched_new = align_blocks(old_blocks, new_blocks, min_sim=args.min_sim)

    write_excel(
        out_path=args.output,
        old_pdf=args.old_pdf,
        new_pdf=args.new_pdf,
        old_blocks=old_blocks,
        new_blocks=new_blocks,
        pairs=pairs,
        unmatched_old=unmatched_old,
        min_sim=args.min_sim,
    )

    print(f"Done. Excel written to: {args.output}")
    if unmatched_new:
        print(f"Info: {len(unmatched_new)} new blocks had no match (reported as 'added').")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
